import requests,re,os,time
from browser import get_merchant_addr_by_asin
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import shutil
from amazoncaptcha import AmazonCaptcha

headers = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36",
    "Cookies": 'session-id=142-2111274-1864942; i18n-prefs=USD; ubid-main=133-5395023-6632938; regStatus=pre-register; aws-target-visitor-id=1680780629969-980708; aws-target-data={"support":"1"}; AMCV_7742037254C95E840A4C98A6@AdobeOrg=1585540135|MCIDTS|19454|MCMID|70846277884120543020010085350038364362|MCAID|NONE|MCOPTOUT-1680787844s|NONE|vVersion|4.4.0; s_fid=2D462374F844CF91-23293BC8ABB62FAE; AMCV_4A8581745834114C0A495E2B@AdobeOrg=-2121179033|MCIDTS|19454|MCMID|41906582696403411915654244310448622955|MCAID|NONE|MCOPTOUT-1680794909s|NONE|vVersion|5.3.0; mbox=PC#b554a32b65594b2890df4865cfd840b7.32_0#1744032510|session#845a61f56ba5485ea287833fca7d9840#1680789570; skin=noskin; session-id-time=2082787201l; sp-cdn="L5Z9:GB"; csm-hit=adb:adblk_yes&t:1680925580498&tb:99B2D08WD2ZZ3345KVX3+s-Q8D1GDV9CZT3XQMZMKRQ|1680925580498; session-token=q38H3/CVLzNJjx6v9G6kItHtbj54FLg8BKKtfGqYL7eWsR91wQXXj5s7psiY00fV8jKB648domicpbrrNAga/sXPM5xI/0Uj7UPSKn1jREKvF0nbkxAwB5i0UfC4RZMtDwpKvL+wxHfPtbeR7Q88kb32M7Gy6bj9AF8ngtGa5ACAfH6cLqSVPTF4xfaSHO9KaSahoFhY+iNV++lJW4ydlcI7ECKGypRUj/05kl5OdhM='
}

# 抓取地址
def get_merchant_addr_by_asin_old(asin):
    global headers
    proxies = {
        'http':'127.0.0.1:7890',
        'https':'127.0.0.1:7890'
        }
    # product_url = f'https://www.amazon.com/DEWALT-DCL040-20-Volt-MAX-Flashlight/dp/B0052MILZM'
    product_url = f'https://www.amazon.com/gp/product/{asin}/address-finder/'  
    print(product_url)
    print('请求ASIN为%s的商品数据' % asin)
    # response = requests.get(url=product_url, headers=headers)
    response = requests.get(url=product_url,proxies=proxies,headers=headers)
    # with open('b.html', 'w',encoding='utf-8') as p:
    #     p.write(response.text)
    soup = BeautifulSoup(response.content, "html.parser")    
    # print('原标题：%s' % len(soup.title.get_text()))
    # 处理验证码
    if len(soup.title.get_text()) < 50:
        print('验证不通过')
    # 卖家数据初始化
    merchant_info = {'asin':asin}
    # 获取卖家id
    merchantID_ele  = soup.find(id='addToCart').find(id='merchantID')
    print(soup.find_all('span',class_='tabular-buybox-text-message'))
    with open('product.html','w',encoding='utf-8') as p:
        p.write(response.text)
    # 最后一个条件是避免有标题正常，但是页面中出现‘We're sorry, an error has occurred. Please reload this page and try again.’这种情况出现
    # print('Please reload this page and try again' in soup.body.get_text())
    # time.sleep(1000)
    if 'orry' in soup.title.get_text() or len(soup.title.string) < 50 or 'Please reload this page and try again' in soup.body.get_text():
        print('可能是Sorry页面，标记为-1，下次重新入选试试；或者可能是有验证码了；或者是商品页面但是出现了Please reload this page and try again')
        merchant_info['name'] = '-1'
        merchant_info['address'] = '-1'
        with open('no_merchantID2.html','w',encoding='utf-8') as p:
            p.write(response.text)
        return merchant_info
    elif len(soup.title.get_text()) > 50 and merchantID_ele is None:
        print('merchantID为空，可能是没商家链接，终止程序前将会保存此网页')
        merchant_info['name'] = 'None'
        merchant_info['address'] = 'None'
        with open('no_merchantID1.html','w',encoding='utf-8') as p:
            p.write(response.text)
        return merchant_info
    merchantID = merchantID_ele.get('value')
    print(merchantID_ele)
    print(merchantID)
    # 拼接商家详情页的url
    merchant_url = f'https://www.amazon.com/gp/help/seller/at-a-glance.html/ref=dp_merchant_link?ie=UTF8&seller={merchantID}&asin={asin}'
    print(merchant_url)
    print('请求商家%s的数据'%merchantID)
    # merchant_res = requests.get(url=merchant_url,headers=headers)
    merchant_res = requests.get(url=merchant_url,proxies=proxies,headers=headers)
    # print(merchant_res.status_code)
    # 处理卖家信息
    merchant_soup = BeautifulSoup(merchant_res.content,'html.parser')
    # 先获取卖家原始数据
    merchant_origin_ele = merchant_soup.find(id='page-section-detail-seller-info')
    if merchant_origin_ele is None:
        merchant_info['name'] = 'None'
        merchant_info['address'] = 'None'
        return merchant_info
    merchant_data_eles = merchant_origin_ele.select('.a-row .a-spacing-none')
    ele = merchant_data_eles[0]
    # 商家信息的标签数量
    span_len = len(ele.select('span'))
    address = ''
    for i in range(span_len):
        if i == 1:
            merchant_info['name'] = ele.select('span')[1].text
        elif i > 2 and i < span_len - 1:
            address = address + ele.select('span')[i].text + '\n'
        elif i == span_len - 1:
            address = address + ele.select('span')[i].text
            merchant_info['address'] = address
    return merchant_info

# 找到与excel中重复的asin，直接将地址等信息复制过来，不再发起网络请求
# 先将重复值一次性写入表格
def need_to_search_filter(asin_list, file_path):
    workbook = load_workbook(filename=file_path)
    # 选择工作表
    worksheet = workbook.active 
    # 获取asin列的值
    asin_col = worksheet['B']
        # print(i)
        # print(i.value)
    after_filter_list = asin_list
    ready_to_write = []
    # 将会删除的位置
    filter_index = -1
    for j in asin_list:
        start_time = time.time()
        filter_index += 1
        # 超过一定数量非期望数据就不再找了，当作他不存在期望数据
        filter_count = 0
        for i in asin_col:
            if j['asin'] == i.value and j['row_index'] > i.row and worksheet['G' + str(i.row)].value is not None:
                # 如果没有正常的数据，继续乡下找
                if len(worksheet['G' + str(i.row)].value) < 5 and filter_count < 2:
                    # print('第%d行的%s，地址不是期望的，继续向下找'%(i.row, i.value))
                    continue
                j['name'] = worksheet['F'+str(i.row)].value
                j['address'] = worksheet['G'+str(i.row)].value
                # print('i:%s,i.index:%d,j:%s,j.index:%d'%(i.value,i.row,j['asin'],j['row_index']))
                # print(type(i.row))
                # print(type(j['row_index']))
                # print('j:%d>i:%d,%s'%(j['row_index'],i.row,j['row_index'] > i.row))
                # print(j)
                del_merchant = after_filter_list.pop(filter_index)
                ready_to_write.append(del_merchant)
                # print('删除了%s'%del_merchant)
                # time.sleep(10)
                break
    end_time1 = time.time()
    now = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(end_time1))
    used_sec = int(end_time1 - start_time)
    print('小步骤1用时%d秒，当前时间%s' %(used_sec, now))
    # 开始写入重复的数据
    start_time2 = time.time()
    # for i in ready_to_write:
    write_list_to_excel(file_path, ready_to_write)
    end_time2 = time.time()
    now = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(end_time2))
    used_sec = int(end_time2 - start_time2)
    print('小步骤2用时%d秒，当前时间%s' %(used_sec, now))
    return after_filter_list


# 获取asin列表
# 需要在这里选出没处理过的asin
def need_to_search(file_path):
    # 用于存放卖家
    merchant_list = []
    workbook = load_workbook(filename=file_path)
    # 选择工作表
    worksheet = workbook.active 
    # 获取asin列的值
    asin_col = worksheet['B']
    # print(worksheet.max_row)
    # print(worksheet.dimensions)
    index = 1
    count = 0
    for cell in asin_col:
        merchant_info = {}
        index+=1
        # 注意这里是根据地址来判断是否要处理此商家
        g_cell = worksheet['G'+str(index)].value
        if g_cell is not None and len(g_cell) > 1:
            # print('%s不需要处理'%(cell.value))
            continue
        count += 1
        if count == 1:
            print('从第%d行开始'% (index))
            print(type(worksheet['F' + str(index)].value))
            print(worksheet['F' + str(index)].value)
        # 装入行数
        merchant_info['row_index'] = index
        # 装入asin
        asin = worksheet['B' + str(index)].value
        merchant_info['asin'] = asin
        if asin is None:
            continue
        # 装入名称
        merchant_info['name'] = worksheet['F' + str(index)].value
        # 装入地址
        merchant_info['address'] = worksheet['G' + str(index)].value
        merchant_list.append(merchant_info)
    # print('原始列表长度:%d' % len(merchant_list))
    merchant_list = need_to_search_filter(merchant_list, file_path)
    # print('过滤后的列表长度:%d' % len(merchant_list))
    return merchant_list

# 写入本地，参数分别是目标文件、卖家数据data_list，data_list是数组，里面存放准备写入的data，而data应该是字典，便于在写入前验证数据
def write_list_to_excel(target_file_path,data_list):
    workbook = load_workbook(filename=target_file_path)
    # 选择工作表
    worksheet = workbook.active 
    print('准备写入数据的长度：%d' % len(data_list))
    # print(data)
    for data in data_list:
        row_index = data['row_index']
        # print('表asin：%s，缓存asin:%s' %(worksheet['B' + str(row_index)].value, data['asin']))
        if worksheet['B' + str(row_index)].value != data['asin']:
            print('数据对不齐了，退出程序')
            return -1
        if len(data['name']) == 0:
            return 0
        worksheet['F' + str(row_index)] = data['name']
        worksheet['G' + str(row_index)] = data['address']
            # row_index += 1
    try:
        workbook.save(filename=target_file_path)
        return 1
    except Exception:
        print('写入失败：程序运行期间不要打开即将生成的目标文件')
        time.sleep(100000)

# 写入本地，参数分别是目标文件、卖家数据data，data应该是字典，便于在写入前验证数据
def write_to_excel(target_file_path,data):
    workbook = load_workbook(filename=target_file_path)
    # 选择工作表
    worksheet = workbook.active 
    print('准备写入以下数据：')
    print(data)
    row_index = data['row_index']
    print('表asin：%s，缓存asin:%s' %(worksheet['B' + str(row_index)].value, data['asin']))
    if worksheet['B' + str(row_index)].value != data['asin']:
        print('数据对不齐了，退出程序')
        return -1
    if len(data['name']) == 0:
        return 0
    worksheet['F' + str(row_index)] = data['name']
    worksheet['G' + str(row_index)] = data['address']
        # row_index += 1
    try:
        workbook.save(filename=target_file_path)
        return 1
    except Exception:
        print('写入失败：程序运行期间不要打开即将生成的目标文件')
        time.sleep(100000)

# 选择要处理的文件
def select_file(origin_dir):
    # 显示当前文件夹的所有子文件夹
    print('当前文件夹：%s' % origin_dir)
    dir_list = os.listdir(origin_dir)
    if len(dir_list) < 1:
        print('文件夹空，先把要处理的文件放到%s文件夹' % origin_dir)
    dir_arr = sorted(dir_list, key=lambda x: os.path.getmtime(os.path.join(origin_dir, x)), reverse=True)
    index = 1
    for name in dir_arr:
        print(index, name)
        index += 1
    print('{0:*^50}'.format('-'))
    print(dir_arr[:6])
    select = int(input('选择要处理的文件（数字序号）：'))
    origin_file = os.path.join(origin_dir, dir_arr[select - 1])
    return origin_file
def prepare_target_file(origin_file_path):
    # 控制程序的进行
    work_flag = 1
    # 先把目标文件复制一份，但在此之前要先判断其是否要是否存在，存在的话是否覆盖
    '''
    if os.path.exists(target_file_path):
        # 因为要捕获输入值为字符时抛出的异常，所以要设置一个循环标志控制循环的继续
        loop_flag = 1
        while loop_flag == 1:
            try:
                print('{0:*^50}'.format('目标文件已存在，是否删除'))
                del_flag = input('【数字1为删除，输入其他数字退出程序】：')
                if int(del_flag) != 1:
                    print('{0:*^50}'.format('退出程序'))
                    print('\n')
                    work_flag = 0
                    break
                else:
                    loop_flag = 0
                    # 删除已存在的目标文件
                    print('原目标文件已删除')
            except Exception:
                print('只能输入数字')
    if work_flag != 1:
        return 0
    '''
    # 目标文件不存在的情况下，复制它，否则继续写入
    if not os.path.exists(target_file_path):
        shutil.copy(origin_file_path, target_file_path)
    return 1

# 如果无法正常跳转，获取验证页面中的图片网址
def get_captcha_img(err_page):
    global headers
    response = requests.get(err_page,headers=headers)
    # with open('c.html', 'w',encoding='utf-8') as p:
    #     p.write(response.text)
    soup = BeautifulSoup(response.text, 'html.parser')
    print('标题：%s' %soup.title.string)
    print('图片数量：%s' % len(soup.select('img')))
    img_link = soup.img['src']
    return img_link

def validate_captcha(img_link):
    captcha = AmazonCaptcha.fromlink(img_link)
    solution = captcha.solve()
    return solution

# if name=='main':
    # validate_captcha()

def main():
    # 获取用户目录
    home_dir = os.path.expanduser('~')
    # 设置工作目录
    origin_dir = os.path.join(home_dir, 'Desktop', 'amazon_task')
    # 目录不存在就创建 
    if not os.path.exists(origin_dir):  
        print('{0:+^80}'.format('工作目录不存在，已创建：【%s】'%origin_dir))
        print('{0:+^50}'.format('先把要处理的Excel文件放到此目录中，再继续运行程序'))
        os.mkdir(origin_dir)
        print('{0:+^50}'.format('按回车继续'))
        input('')
    origin_file = select_file(origin_dir=origin_dir)
    print('源文件：%s' % origin_file)
    merchant_list = []
    # 循环开始时间，用于计算每次循环的耗时
    start_time = time.time()
    # 先读取excel，获得要处理的卖家列表，每个元素是个字典，其应包含：row_index,asin,name,address
    meta_data = need_to_search(file_path=origin_file)
    # time.sleep(3)
    # print(meta_data)
    return meta_data


if __name__=='__main__':
    # main()
    strr = 'C:/Users/x/Desktop/amazon_task/自行车灯 - 副本 (2).xlsx'
    meta_data = need_to_search(file_path=strr)
    # print(meta_data)
    # need_to_search2(meta_data, file_path=strr)